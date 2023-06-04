from openpyxl import load_workbook
ws = load_workbook(filename = 'Reisen - v2.xlsx')
database = dict()
#print(sheet_ranges['B2'].value)
for row in range(10):
    print(row)
    id = ws.cell(row,10)
    entry = ws.cell(row, 11)
    exit = ws.cell(row,12)
    database = {id: entry}
    print(id)
print(database)